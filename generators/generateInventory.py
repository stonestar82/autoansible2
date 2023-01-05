from multiprocessing.sharedctypes import Value
import re, ipaddress
import yaml, json
from generators.BlankNone import *
from openpyxl import load_workbook
from operator import eq
import pandas as pd
from jinja2 import Template
from PIL import Image, ImageFont, ImageDraw

def convertToBoolIfNeeded(variable):
	if type(variable) == str and re.match(r'(?i)(True|False)', variable.strip()):
		variable = True if re.match(r'(?i)true', variable.strip()) else False
	return variable

def getFabricName(inventory_file, excelVar):
  workbook = load_workbook(filename=inventory_file, read_only=False, data_only=True)
  return getExcelSheetValue(workbook, excelVar["all"]["fabricName"])

def parseSpineInfo(inventory_file, excelVar):
	'''
	엑셀에서 데이터를 읽어 spine 정보 처리
	'''
	spines_info = {"vars": {"type": "spine"}, "hosts": {}}
	workbook = load_workbook(filename=inventory_file, read_only=True, data_only=True)
	inventory_worksheet = workbook[excelVar["spine"]["sheet"]]

	spinePrefix = excelVar["spine"]["prefix"]
	spineHostnameCol = excelVar["spine"]["props"]["hostname"]["col"]

	for row in inventory_worksheet.iter_rows():
		for cell in row:
			# print(cell.value)
			if cell.value:
				if eq(cell.coordinate, spineHostnameCol + str(cell.row)):
					p = re.compile(spinePrefix)
					if (p.match(str(cell.value))):
						codi = excelVar["spine"]["props"]["mgmt"]["col"] + str(cell.row)
						mgmtIp = inventory_worksheet[codi].value
						spines_info["hosts"][cell.value] = {"ansible_host": mgmtIp}
	
	return spines_info

def parseLeafInfo(inventory_file, excelVar, leaf_type="L3"):
	'''
	엑셀에서 데이터를 읽어 leaf 정보 처리
	'''
	
	workbook = load_workbook(filename=inventory_file, read_only=True, data_only=True)
	leafTypeName = "l3leaf" if leaf_type == "L3" else "l2leaf"
	inventory_worksheet = workbook[excelVar["leaf"]["sheet"]]
	leafs = {"vars": {"type": leafTypeName}, "hosts": {}}
	
	# transform the workbook to a list of dictionaries
	leafPrefix = excelVar["leaf"]["prefix"]
	leafHostnameCol = excelVar["leaf"]["props"]["hostname"]["col"]
	
	for row in inventory_worksheet.iter_rows():
		for cell in row:
			# print(cell)
			if cell.value:
				if eq(cell.coordinate, leafHostnameCol + str(cell.row)):
					p = re.compile(leafPrefix)
					if (p.match(str(cell.value))):
						codi = excelVar["leaf"]["props"]["mgmt"]["col"] + str(cell.row)
						mgmtIp = inventory_worksheet[codi].value
      
						leafs["hosts"][cell.value] = {"ansible_host": mgmtIp}


	return leafs

def getFabricInventory(inventory_file, fabric_name, excelVar):
	"""
	엑셀에서 데이터를 읽어 inventory.yml 생성용 데이터 처리
	"""
	fabric_inventory = {"children":{}}
	workbook = load_workbook(filename=inventory_file, read_only=False, data_only=True)

	fabric_inventory["children"][fabric_name+"_SPINES"] = parseSpineInfo(inventory_file, excelVar)
	
	if parseLeafInfo(inventory_file, excelVar, leaf_type="L3") != None:
		fabric_inventory["children"][fabric_name+"_L3LEAFS"] = parseLeafInfo(inventory_file, excelVar, leaf_type="L3")
		
	fabric_inventory["vars"] = {
		"ansible_connection": "network_cli",
		"ansible_network_os": "eos",
		"ansible_become": True,
		"ansible_user": getExcelSheetValue(workbook, excelVar["all"]["ansibleName"]),
    "ansible_ssh_pass": getExcelSheetValue(workbook, excelVar["all"]["ansiblePassword"]),
		"ansible_become_method": "enable",
		"ansible_httpapi_use_ssl": False,
		"ansible_httpapi_validate_certs": False
	}
	return fabric_inventory

def generateInventory(inventory_file, excelVar):
	"""
	엑셀에서 데이터를 읽어 inventory 정보 처리
	d1.yml, all.yml 파일 생성
  toplogy 이미지 생성
	"""
	fabric_name = getFabricName(inventory_file, excelVar)

	if fabric_name is None:
		return

	inventory = {
		"all": {
			"children": {
				"FABRIC": {
					"children": {
						fabric_name: {
							"children": {       
								"PODS": {
									"children": { 
										fabric_name + "_SPINES" : None,
										fabric_name + "_L3LEAFS" : None
									}
								}
							}
						}
					}
				}
			}
		}
	}

	workbook = load_workbook(filename=inventory_file, read_only=False, data_only=True)
 
	info = {}
	for item in excelVar["all"]:
		v = getExcelSheetValue(workbook, excelVar["all"][item])
		info[excelVar["all"][item]["mapping"]] = v

	mgmtVrf = info["mgmt_interface_vrf"]
	mgmtInterface = info["mgmt_interface"]
	mgmtGw = info["mgmt_gateway"]
	macAging = info["mac_aging"]
	arpAging = info["arp_aging"]
	timeZone = info["time_zone"]
	adminName = info["admin_name"]
	adminPassword = info["admin_info"]
	admin_privilege = info["admin_privilege"]
	spanningTreeMode = info["spanning_tree_mode"]
	terminalLength = info["terminal_length"]
	terminalWidth = info["terminal_width"]
	logginBufferedLevel = info["loggin_buffered_level"]
	p2pSubnet = info["p2p_subnet"]
 
	#Add Fabric info
	inventory["all"]["children"]["FABRIC"]["children"][fabric_name]["children"]["PODS"] = getFabricInventory(inventory_file, fabric_name, excelVar)

	##### port map 정보 로드 S #####
	sheetName = excelVar["pd"]["portMap"]["sheetName"]
	headerRow = excelVar["pd"]["portMap"]["header"]
	spineCol = excelVar["pd"]["portMap"]["spine"]
	spinePortCol = excelVar["pd"]["portMap"]["spinePort"]
	spineIpCol = excelVar["pd"]["portMap"]["spineIp"]
	leafCol = excelVar["pd"]["portMap"]["leaf"]
	leafPortCol = excelVar["pd"]["portMap"]["leafPort"]
	leafIpCol = excelVar["pd"]["portMap"]["leafIp"]

	spinePrefix = excelVar["spine"]["prefix"]
	leafPrefix = excelVar["leaf"]["prefix"]

	switches = pd.read_excel(inventory_file, header=headerRow, sheet_name=sheetName)[[spineCol, spinePortCol, spineIpCol, leafCol, leafPortCol, leafIpCol]].dropna(axis=0)

	portMap = {}

	for idx, switch in switches.iterrows():
		
		##### P2P S #####  
		## spine switch 정리
		p = re.compile(spinePrefix)
		if (p.match(str(switch[spineCol]))):
			spine = switch[spineCol]
			if not spine in portMap:
				portMap.setdefault(
					spine,  {
						"INTERFACES": [{ "ETHERNET": switch[spinePortCol], "IP": switch[spineIpCol] }],
						"ETC_PORTS": { "IP": "", "INTERFACES": [] }
					}
				)
			else:
				portMap[spine]["INTERFACES"].append({ "ETHERNET": switch[spinePortCol], "IP": switch[spineIpCol] })
				
		## leaf switch 정리
		p = re.compile(leafPrefix)
		leaf = switch[spineCol]
		if (p.match(str(switch[leafCol])) and not p.match(leaf)):
			leaf = switch[leafCol]
			if not leaf in portMap:
				portMap.setdefault(
					leaf,  {
						"INTERFACES": [{ "ETHERNET": switch[leafPortCol], "IP": switch[leafIpCol] }],
						"ETC_PORTS": { "IP": "", "INTERFACES": [] }
					}
				)
			else:
				portMap[leaf]["INTERFACES"].append({ "ETHERNET": switch[leafPortCol], "IP": switch[leafIpCol] })
    ##### P2P E ##### 
    
    ##### port channell S #####
		p = re.compile(leafPrefix)
		# print(switch[spineCol], switch[leafCol])
		if (p.match(str(switch[leafCol])) and p.match(str(switch[spineCol]))):

			leaf = switch[leafCol]
			portMap[leaf]["ETC_PORTS"]["IP"] = switch[leafIpCol]
			portMap[leaf]["ETC_PORTS"]["INTERFACES"].append({ "ETHERNET": switch[leafPortCol] })

			leaf = switch[spineCol]
			portMap[leaf]["ETC_PORTS"]["IP"] = switch[spineIpCol]
			portMap[leaf]["ETC_PORTS"]["INTERFACES"].append({ "ETHERNET": switch[spinePortCol] })
   
		##### port channell E #####	
    
    
	##### port map 정보 로드 E #####

	## 기본변수 로드
	with open("./excelEnvriment.json", "r", encoding='utf8') as f:
		excelVar = json.load(f)
		f.close()
  
	sheetName = excelVar["pd"]["switchIpInfo"]["sheetName"]
	headerRow = excelVar["pd"]["switchIpInfo"]["header"]
	hostNameCol = excelVar["pd"]["switchIpInfo"]["hostName"]
	mgmtCol = excelVar["pd"]["switchIpInfo"]["mgmt"]
	loopback0Col = excelVar["pd"]["switchIpInfo"]["loopback0"]
	bgpAsnCol = excelVar["pd"]["switchIpInfo"]["bgpAsn"]
	typeCol = excelVar["pd"]["switchIpInfo"]["type"]
	idCol = excelVar["pd"]["switchIpInfo"]["id"]
	loop1Col = excelVar["pd"]["switchIpInfo"]["loopback1"]
	
 
	## Switch 정보 로드
	switches = pd.read_excel(inventory_file, header=headerRow, sheet_name=sheetName)[[hostNameCol, mgmtCol, loopback0Col, bgpAsnCol, typeCol, idCol, loop1Col]].fillna("")

	config = {"hosts": None}

	data = {}
	## spine, leaf, bl 개수 체크
	topologySwitches = { "spine": 0, "leaf": 0, "bl": 0 }

	for idx, switch in switches.iterrows():
		hostname = switch[hostNameCol]
		mgmt = switch[mgmtCol]
		loop0 = switch[loopback0Col]
		
		data.setdefault(
			hostname,  {
					"HOSTNAME": hostname,
					"HOST_IP": mgmt,
					"LOOPBACK0": loop0,
					"LOOPBACK1": str(switch[loop1Col]) + "/32" if switch[loop1Col] != "" else "",
					"PERMIT_IP": str(ipaddress.IPv4Interface(str(switch[loop1Col]) + "/24").network) if switch[loop1Col] != "" else "",					
					"INTERFACES": portMap[hostname]["INTERFACES"],
					"ETC_PORTS": portMap[hostname]["ETC_PORTS"],
					"BGP_ASN": int(switch[bgpAsnCol]) if switch[typeCol] == "Spine" else switch[bgpAsnCol],
					"ID": switch[idCol],
					"TYPE": switch[typeCol],
					"P2P_SUBNET": p2pSubnet
			}
		)
  
		## spine, leaf, bl 갯수 체크
		v = topologySwitches[str(switch[typeCol]).lower()]
		topologySwitches[str(switch[typeCol]).lower()] = v + 1
   
  
  
	config["hosts"] = data

	with BlankNone(), open("./inventory/group_vars/" + fabric_name + ".yml", "w") as inv:
		inv.write(yaml.dump(config, sort_keys=False))
		inv.close()
  
  # Group Vars all.yml 파일 생성
	data = {
		"TERMINAL_LENGTH": terminalLength,
		"TERMINAL_WIDTH": terminalWidth,
		"LOGGIN_BUFFERED": logginBufferedLevel,
		"SPANNING_TREE_MODE": spanningTreeMode,
		"ADMIN_USER_NAME": adminName,
		"ADMIN_USER_PW": adminPassword,
		"TIMEZONE": timeZone,
		"ARP_AGING": arpAging,
		"MAC_AGING": macAging,
		"ADMIN_PRIVILEGE": admin_privilege,
		"MGMT_VRF": mgmtVrf,
		"MGMT_INTERFACE": mgmtInterface,
		"MGMT_GW": mgmtGw,
		"BACKUP_FILENAME": "{{ inventory_hostname }}_{{ lookup('pipe', 'date +%Y%m%d%H%M%S') }}"
	}

	with open('./inventory/templates/playbook/allyml.j2') as f:
		template = Template(f.read())

	with open("./inventory/group_vars/all.yml", "w") as reqs:
			reqs.write(template.render(**data))


	##### toplogy 이미지 생성 S #####
	# topologySwitches = { "spine": 2, "leaf": 3, "bl": 0 }
	spinesCount = topologySwitches["spine"]
	leafsCount = topologySwitches["leaf"]
	blsCount = topologySwitches["bl"]
	totalLeafsCount = leafsCount + blsCount
	width = 1920
	height = 1080
	space = 200 ## switch icon 사이 간격
	bgcolor = "#28353c"
	spineTopMargin = 200
	leafTopMargin = 600
	iconWidth = 50
 
	## spine icon 시작 위치값
	s_start = int(width / spinesCount - space / spinesCount)
	if eq(1, spinesCount):
		s_start = int(width / 2 - (iconWidth / 2))
  
  ## leaf icon 시작 위치값
	l_start = int(width / totalLeafsCount + space - (iconWidth/2))

	canvas = Image.new("RGB", (width, height), bgcolor)
	font = ImageFont.truetype("./image/verdanab.ttf", 13)

	for i in range (spinesCount):	
		img = Image.open("./image/switch.png")
		draw = ImageDraw.Draw(img)
		text = fabric_name + "-Spine" + str(i + 1)
		draw.text((5, 72), text, font=font)
		
		canvas.paste(img, (s_start + (i * space), spineTopMargin))

	p = 0
	imgName = "Leaf"

	for i in range(totalLeafsCount):
		p = p + 1
  
		if (p > leafsCount):
			imgName = "BL"
			p = 1

		img = Image.open("./image/switch.png")
		draw = ImageDraw.Draw(img)
		text = fabric_name + "-" + imgName + str(i + 1)
		draw.text((5, 72), text, font=font)
		canvas.paste(img, (l_start + (i * space), leafTopMargin))

	canvas.save("topology.png")
	##### toplogy 이미지 생성 E #####

	return inventory